"""세금계산서 발행 일정표 xlsx 파싱. 기존 mailing/src/xlsxParser.ts 를 이식.

연도별 섹션이 반복(각 섹션마다 헤더 + 소계/합계 행)되는 표. 헤더 행을 만날 때마다
컬럼 매핑을 다시 잡고, 소계/합계 행은 건너뛴다. 병합 셀(사업구분/거래처)은 위 값을 이어받는다.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import load_workbook

HEADER_MARKERS = ("거래처", "발행 예정일")
MONTH_HEADERS = ["1월", "2월", "3월", "4월", "5월", "6월",
                 "7월", "8월", "9월", "10월", "11월", "12월"]
STOP_MARKERS = ("소계", "합계")


@dataclass
class InvoiceRow:
    business_type: str
    client_name: str
    maintenance_detail: str
    issue_status: str
    schedule_pattern: str
    ceo_name: str
    contact_name: str
    contact_email: str
    contact_phone: str
    sales_type: str
    monthly_amount: int | None


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def parse_invoice_rows(data: bytes, target_month: int) -> list[InvoiceRow]:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    results: list[InvoiceRow] = []
    header_cols: dict[str, int] | None = None
    last_business = ""
    last_client = ""

    for row in ws.iter_rows(values_only=True):
        cells = [_norm(c) for c in row]

        if all(m in cells for m in HEADER_MARKERS):
            header_cols = {c: i for i, c in enumerate(cells) if c}
            last_business = ""
            last_client = ""
            continue
        if not header_cols:
            continue

        def get(name: str) -> str:
            idx = header_cols.get(name)
            if idx is None or idx >= len(cells):
                return ""
            return cells[idx]

        client_raw = get("거래처")
        business_raw = get("사업 구분")
        if client_raw in STOP_MARKERS or business_raw in STOP_MARKERS:
            continue

        business = business_raw or last_business
        client = client_raw or last_client
        last_business, last_client = business, client

        detail = get("유지보수 내역")
        if not client or not detail:
            continue

        month_col = MONTH_HEADERS[target_month - 1]
        raw_amount = get(month_col)
        amount = round(float(raw_amount)) if raw_amount else None

        results.append(InvoiceRow(
            business_type=business,
            client_name=client,
            maintenance_detail=detail,
            issue_status=get("발행 상태"),
            schedule_pattern=get("발행 예정일"),
            ceo_name=get("대표자"),
            contact_name=get("발행담당자명"),
            contact_email=get("발행담당자이메일"),
            contact_phone=get("발행담당자연락처"),
            sales_type=get("매출형태"),
            monthly_amount=amount,
        ))
    return results
